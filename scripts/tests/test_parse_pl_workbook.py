#!/usr/bin/env python3
"""Checks for the P&L workbook parser, run as `python3 scripts/tests/test_...py`.

Fixtures are synthetic. No real revenue, expense, or personal value appears here, and
none may be added: this repository is public.
"""

from __future__ import annotations

import json
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
PARSER = ROOT / "scripts" / "parse_pl_workbook.py"

# A minimal but structurally faithful sheet: three revenue categories with their VAT
# lines, four weeks, a two-stream split, four expense groups, one of which has no lines.
BASE = {
    "weeks": [40, 30, 20, 10],
    "streams": [("2F", 70), ("1F", 30)],
    "categories": [("Beverage", 50, 5), ("Food", 40, 4), ("Service charge", 1, 0)],
    "groups": [
        ("1 - Rental fee", 20, []),
        ("2 - Payroll", 30, [("Salary", 25), ("Insurance", 5)]),
        ("3 - Cost of sales", 10, [("Beverage", 6), ("Food", 4)]),
        ("4 - Operating", 5, [("Electricity", 3), ("Other", 2)]),
    ],
}


def build(path: Path, spec: dict) -> None:
    """Write a workbook shaped like the real one, from a spec of period-level figures."""
    weeks = spec["weeks"]
    share = [w / sum(weeks) for w in weeks]
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "P&L"
    sheet.append([])
    header = [None, "Actual Monthly", "Week 1", "Week 2", "Week 3", "Week 4", None, None, None]
    header[7], header[8] = spec["streams"][0][0], spec["streams"][1][0]
    sheet.append(header)

    revenue = sum(ex + vat for _, ex, vat in spec["categories"])
    row = [None] * 9
    row[0], row[1] = "I - Net Revenue (A-B)", revenue
    for index, value in enumerate(weeks):
        row[2 + index] = value
    row[7], row[8] = spec["streams"][0][1], spec["streams"][1][1]
    sheet.append(row)
    for name, ex, vat in spec["categories"]:
        sheet.append([name, ex, *[ex * s for s in share]])
        sheet.append([f"{name} - output VAT", vat, *[vat * s for s in share]])

    expenses = sum(total for _, total, _ in spec["groups"])
    sheet.append(["II - Expenses", expenses, *[expenses * s for s in share]])
    for name, total, lines in spec["groups"]:
        sheet.append([name, total, *[total * s for s in share]])
        for line_name, line_total in lines:
            sheet.append([line_name, line_total, *[line_total * s for s in share]])
        sheet.append([0, 0, 0, 0, 0])  # residual row, as the real sheet carries
    sheet.append(["III - Profit (loss)", revenue - expenses])
    book.save(path)


def build_sales(path: Path, amounts: list[float], stated: float | None) -> None:
    """Write a workbook shaped like the POS sales export: a preamble, a header, rows,
    then a row of column totals."""
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["a preamble line the export always carries"])
    sheet.append(["店铺", "项目名", "总金额（含佣金）", "总金额"])
    for index, value in enumerate(amounts):
        sheet.append([f"shop", f"item {index}", value, value])
    if stated is not None:
        sheet.append(["Tổng", "-", stated, stated])
    book.save(path)


def run(
    path: Path,
    out: Path | None = None,
    verify: Path | None = None,
    stream: str | None = None,
) -> subprocess.CompletedProcess[str]:
    command = [
        sys.executable, str(PARSER), str(path),
        "--client-id", "TEST", "--project-id", "TEST",
        "--period-start", "2026-08-01", "--period-end", "2026-08-31",
        "--currency", "XTS",
    ]
    if out:
        command += ["--out", str(out)]
    if verify:
        command += ["--verify-sales", str(verify)]
    if stream:
        command += ["--verify-stream", stream]
    return subprocess.run(command, capture_output=True, text=True)


def check(name: str, condition: bool, detail: str = "") -> bool:
    print(f"  {'PASS' if condition else 'FAIL'}  {name}{'  ' + detail if detail else ''}")
    return condition


def main() -> int:
    results: list[bool] = []
    with tempfile.TemporaryDirectory() as tmp:
        work = Path(tmp)

        print("a well-formed workbook is accepted")
        book, out = work / "ok.xlsx", work / "ok.json"
        build(book, BASE)
        result = run(book, out)
        results.append(check("exit status is 0", result.returncode == 0, result.stderr.strip()))
        if result.returncode == 0:
            payload = json.loads(out.read_text())
            revenue = payload["revenue_records"]
            total = next(r for r in revenue if r["breakdown"] == "TOTAL")
            results.append(check("TOTAL equals the category sum", total["amount_incl_vat"] == 100))
            streams = [r for r in revenue if r["breakdown"] == "STREAM"]
            results.append(check("both streams are recorded", len(streams) == 2))
            results.append(
                check(
                    "stream records carry no invented VAT split",
                    all(r["amount_ex_vat"] is None and r["output_vat"] is None for r in streams),
                )
            )
            groups = [
                r for r in payload["expense_records"]
                if r["breakdown"] == "GROUP" and r["scope"] is None
            ]
            results.append(check("all four expense groups are recorded", len(groups) == 4))
            results.append(
                check(
                    "a group stated without lines is kept",
                    any(g["group_code"] == "1" for g in groups),
                )
            )
            results.append(
                check("residual zero rows do not become records",
                      not any(r.get("line_label") == "0" for r in payload["expense_records"]))
            )
            results.append(check("the source digest is recorded", len(payload["source_sha256"]) == 64))

        print("\na workbook whose streams do not sum to the total is rejected")
        broken = dict(BASE, streams=[("2F", 70), ("1F", 25)])
        book, out = work / "bad_stream.xlsx", work / "bad_stream.json"
        build(book, broken)
        result = run(book, out)
        results.append(check("exit status is 1", result.returncode == 1))
        results.append(check("the failing rule is named", "REVENUE_STREAM_SUM" in result.stderr))
        results.append(check("nothing is written", not out.exists()))

        print("\na workbook whose expense lines do not sum to their group is rejected")
        groups = list(BASE["groups"])
        groups[1] = ("2 - Payroll", 30, [("Salary", 25), ("Insurance", 4)])
        book = work / "bad_lines.xlsx"
        build(book, dict(BASE, groups=groups))
        result = run(book)
        results.append(check("exit status is 1", result.returncode == 1))
        results.append(check("the failing rule is named", "EXPENSE_LINE_SUM" in result.stderr))

        print("\na sales export matching the stream is recorded as MATCHED")
        book, out = work / "verify_ok.xlsx", work / "verify_ok.json"
        build(book, BASE)
        sales = work / "sales_ok.xlsx"
        build_sales(sales, [40.0, 30.0], 70.0)  # equals the 2F stream figure
        result = run(book, out, verify=sales, stream="2F")
        results.append(check("exit status is 0", result.returncode == 0, result.stderr.strip()))
        if out.exists():
            verification = json.loads(out.read_text())["verifications"][0]
            results.append(check("status is MATCHED", verification["status"] == "MATCHED"))
            results.append(check("difference is zero", verification["difference"] == 0))
            results.append(
                check("the compared column is named", "总金额" in verification["external_source"])
            )

        print("\na sales export that disagrees is DIVERGED, and the period is still accepted")
        book, out = work / "verify_diff.xlsx", work / "verify_diff.json"
        build(book, BASE)
        sales = work / "sales_diff.xlsx"
        build_sales(sales, [40.0, 25.0], 65.0)
        result = run(book, out, verify=sales, stream="2F")
        results.append(check("exit status is 0", result.returncode == 0))
        results.append(check("a warning is printed", "DIVERGED" in result.stderr))
        if out.exists():
            verification = json.loads(out.read_text())["verifications"][0]
            results.append(check("status is DIVERGED", verification["status"] == "DIVERGED"))
            results.append(check("the difference is recorded", verification["difference"] == 5))

        print("\na sales export inconsistent with its own total row is rejected")
        book = work / "verify_trunc.xlsx"
        build(book, BASE)
        sales = work / "sales_trunc.xlsx"
        build_sales(sales, [40.0], 70.0)  # a row is missing
        result = run(book, verify=sales, stream="2F")
        results.append(check("exit status is 1", result.returncode == 1))
        results.append(
            check("the reason is named", "SALES_EXPORT_INCONSISTENT" in result.stderr)
        )

        print("\nverifying a stream the workbook does not have is rejected")
        book = work / "verify_stream.xlsx"
        build(book, BASE)
        sales = work / "sales_stream.xlsx"
        build_sales(sales, [70.0], 70.0)
        result = run(book, verify=sales, stream="3F")
        results.append(check("exit status is 1", result.returncode == 1))
        results.append(check("the available streams are listed", "'2F'" in result.stderr))

        print("\na workbook with no P&L sheet is rejected")
        book = work / "no_sheet.xlsx"
        empty = openpyxl.Workbook()
        empty.active.title = "Sheet1"
        empty.save(book)
        result = run(book)
        results.append(check("exit status is 1", result.returncode == 1))
        results.append(check("the reason is named", "no 'P&L' sheet" in result.stderr))

    print(f"\n{sum(results)}/{len(results)} checks passed")
    return 0 if all(results) else 1


if __name__ == "__main__":
    sys.exit(main())
