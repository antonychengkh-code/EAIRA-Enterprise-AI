#!/usr/bin/env python3
"""Parse the `P&L` sheet of a monthly accounting workbook into a period payload.

Emits the request body defined by `docs/finance/backend/revenue_input_api.md`, after
checking the identities the schemas require. A workbook that fails any identity is
rejected whole: the payload is not written and the exit status is non-zero. Partial
output would be worse than none, because a period that is internally inconsistent looks
valid one record at a time.

Only the `P&L` sheet is read. The weekly sheets are maintained by hand and their layout
moves between sheets and between months.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import re
import sys
from pathlib import Path
from typing import Any, Iterable

import openpyxl

SHEET = "P&L"
REVENUE_HEAD = "I - "
EXPENSE_HEAD = "II - "
PROFIT_HEAD = "III - "
GROUP_LABEL = re.compile(r"^(\d+)\s*-\s*(.+)$")
VAT_SUFFIX = re.compile(r"\s*-\s*(output|input)\s+VAT$", re.IGNORECASE)

MONTH_COL = 1
WEEK_COLS = (2, 3, 4, 5)
STREAM_COLS = (7, 8)

# The sales export's gross-amount column, as titled by each locale the export offers.
# Matched exactly: neighbouring columns contain this text as a substring.
D05_TOTAL_HEADERS = ("总金额", "總金額", "Tổng tiền", "Total amount")
# The export closes with a row of column totals, labelled in the first column.
D05_TOTAL_MARKERS = ("tổng", "总计", "总额", "總計", "total")
# The export's heading states the window it covers, and so states its own business day
# cutoff. Read it from there: a cutoff supplied by hand would not be evidence.
D05_WINDOW = re.compile(
    r"(\d{2})/(\d{2})/(\d{4})\s+(\d{2}:\d{2}).*?(\d{2})/(\d{2})/(\d{4})\s+(\d{2}:\d{2})"
)

AUTHORITATIVE = "AUTHORITATIVE"
VERIFICATION = "VERIFICATION"


class Rejected(Exception):
    """A workbook that violates the schemas' identities."""

    def __init__(self, failures: list[str]) -> None:
        super().__init__("; ".join(failures))
        self.failures = failures


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def label_of(row: Iterable[Any]) -> str:
    """Return a row's line label, or an empty string when the row carries none.

    Residual rows in the sheet hold a numeric zero where a label would be. They are not
    lines and must not become records.
    """
    first = next(iter(row), None)
    return first.strip() if isinstance(first, str) else ""


def amount(row: list[Any], column: int) -> float | None:
    if column >= len(row):
        return None
    value = row[column]
    return float(value) if isinstance(value, (int, float)) else None


def read_pl(path: Path) -> list[list[Any]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET not in workbook.sheetnames:
        raise Rejected([f"workbook has no {SHEET!r} sheet"])
    return [list(row) for row in workbook[SHEET].iter_rows(values_only=True)]


def section_bounds(rows: list[list[Any]]) -> tuple[int, int, int]:
    """Locate the revenue, expense, and profit heading rows."""
    found: dict[str, int] = {}
    for index, row in enumerate(rows):
        text = label_of(row)
        for head in (REVENUE_HEAD, EXPENSE_HEAD, PROFIT_HEAD):
            if head not in found and text.startswith(head):
                found[head] = index
    missing = [h.strip(" -") for h in (REVENUE_HEAD, EXPENSE_HEAD, PROFIT_HEAD) if h not in found]
    if missing:
        raise Rejected([f"missing section heading {name!r}" for name in missing])
    return found[REVENUE_HEAD], found[EXPENSE_HEAD], found[PROFIT_HEAD]


def split_vat(label: str) -> tuple[str, bool]:
    """Split a category label into its base name and whether it is the VAT line."""
    stripped = VAT_SUFFIX.sub("", label)
    return stripped.strip(), stripped != label


def build_revenue(
    rows: list[list[Any]], revenue_row: int, expense_row: int, week_labels: list[str]
) -> list[dict[str, Any]]:
    """Build revenue records for every breakdown the sheet actually provides."""
    head = rows[revenue_row]
    records: list[dict[str, Any]] = []

    # Categories carry their own VAT line, so pair each base name with its VAT row.
    net: dict[str, dict[str, float | None]] = {}
    order: list[str] = []
    for row in rows[revenue_row + 1 : expense_row]:
        label = label_of(row)
        if not label:
            continue
        base, is_vat = split_vat(label)
        entry = net.setdefault(base, {"ex_vat": None, "vat": None})
        if base not in order:
            order.append(base)
        key = "vat" if is_vat else "ex_vat"
        entry[key] = amount(row, MONTH_COL)
        for column, week in zip(WEEK_COLS, week_labels):
            entry.setdefault(f"{key}_{week}", amount(row, column))  # type: ignore[index]

    total_ex = sum(v["ex_vat"] or 0.0 for v in net.values())
    total_vat = sum(v["vat"] or 0.0 for v in net.values())
    records.append(
        {
            "breakdown": "TOTAL",
            "scope": None,
            "category": None,
            "stream": None,
            "amount_ex_vat": total_ex,
            "output_vat": total_vat,
            "amount_incl_vat": total_ex + total_vat,
        }
    )
    for base in order:
        entry = net[base]
        ex_vat = entry["ex_vat"] or 0.0
        vat = entry["vat"] or 0.0
        records.append(
            {
                "breakdown": "CATEGORY",
                "scope": None,
                "category": base,
                "stream": None,
                "amount_ex_vat": ex_vat,
                "output_vat": vat,
                "amount_incl_vat": ex_vat + vat,
            }
        )
    for column, week in zip(WEEK_COLS, week_labels):
        ex_vat = sum(entry.get(f"ex_vat_{week}") or 0.0 for entry in net.values())
        vat = sum(entry.get(f"vat_{week}") or 0.0 for entry in net.values())
        records.append(
            {
                "breakdown": "WEEK",
                "scope": week,
                "category": None,
                "stream": None,
                "amount_ex_vat": ex_vat,
                "output_vat": vat,
                "amount_incl_vat": ex_vat + vat,
            }
        )
    # The stream split exists on the heading row only, and without a VAT breakdown.
    for column in STREAM_COLS:
        stream = rows[revenue_row - 1][column] if revenue_row else None
        value = amount(head, column)
        if not isinstance(stream, str) or value is None:
            continue
        records.append(
            {
                "breakdown": "STREAM",
                "scope": None,
                "category": None,
                "stream": stream.strip(),
                "amount_ex_vat": None,
                "output_vat": None,
                "amount_incl_vat": value,
            }
        )
    return records


def build_expenses(
    rows: list[list[Any]], expense_row: int, profit_row: int, week_labels: list[str]
) -> list[dict[str, Any]]:
    """Build expense records, preserving the group and line hierarchy."""
    records: list[dict[str, Any]] = [
        {
            "breakdown": "TOTAL",
            "scope": None,
            "group_code": None,
            "group_label": None,
            "line_label": None,
            "amount": amount(rows[expense_row], MONTH_COL) or 0.0,
        }
    ]
    for column, week in zip(WEEK_COLS, week_labels):
        value = amount(rows[expense_row], column)
        if value is not None:
            records.append(
                {
                    "breakdown": "TOTAL",
                    "scope": week,
                    "group_code": None,
                    "group_label": None,
                    "line_label": None,
                    "amount": value,
                }
            )

    group_code = group_label = None
    for row in rows[expense_row + 1 : profit_row]:
        label = label_of(row)
        if not label:
            continue
        match = GROUP_LABEL.match(label)
        breakdown = "GROUP" if match else "LINE"
        if match:
            group_code, group_label = match.group(1), match.group(2).strip()
        elif group_code is None:
            raise Rejected([f"expense line {label!r} appears before any group"])
        common = {
            "group_code": group_code,
            "group_label": group_label,
            "line_label": None if match else label,
        }
        for column, week in [(MONTH_COL, None), *zip(WEEK_COLS, week_labels)]:
            value = amount(row, column)
            if value is None:
                continue
            records.append({"breakdown": breakdown, "scope": week, **common, "amount": value})
    return records


def declared_window(rows: list[list[Any]]) -> tuple[str, str] | None:
    """Return the window a sales export states in its heading, as (start date, cutoff)."""
    for row in rows[:3]:
        for cell in row:
            if not isinstance(cell, str):
                continue
            found = D05_WINDOW.search(cell)
            if found:
                day, month, year, cutoff = found.group(1, 2, 3, 4)
                return f"{year}-{month}-{day}", cutoff
    return None


def scan_sales_sheet(rows: list[list[Any]]) -> tuple[float, float | None, str] | None:
    """Read one sheet of a sales export, or return None when it is not a usable one."""
    header_index = column = None
    for index, row in enumerate(rows[:10]):
        for position, cell in enumerate(row):
            if isinstance(cell, str) and cell.strip() in D05_TOTAL_HEADERS:
                header_index, column = index, position
                break
        if column is not None:
            break
    if column is None:
        return None

    stated: float | None = None
    data_sum = 0.0
    for row in rows[header_index + 1 :]:
        first = row[0] if row else None
        value = row[column] if column < len(row) else None
        if not isinstance(value, (int, float)):
            continue
        if isinstance(first, str) and first.strip().lower() in D05_TOTAL_MARKERS:
            stated = float(value)
        else:
            data_sum += float(value)
    return data_sum, stated, str(rows[header_index][column]).strip()


def read_sales_total(path: Path, sheet: str | None = None) -> tuple[float, str, str | None, str | None]:
    """Return the sales export's gross total, checked against its own stated total.

    The export ends in a row repeating each column's sum in the data columns. Summing
    the data rows and comparing against that stated total catches a truncated or edited
    export, which would otherwise verify against nothing.

    An export may hold a per-store sheet alongside an all-stores sheet, and only the
    latter labels its total row. Sheets are therefore tried in order and the first
    usable one is taken, unless one is named explicitly.
    """
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = [sheet] if sheet else workbook.sheetnames
    if sheet and sheet not in workbook.sheetnames:
        raise Rejected([f"{path.name}: no sheet named {sheet!r}; it has {workbook.sheetnames}"])

    seen_column = False
    for name in names:
        rows = [list(r) for r in workbook[name].iter_rows(values_only=True)]
        found = scan_sales_sheet(rows)
        if found is None:
            continue
        data_sum, stated, header_name = found
        seen_column = True
        if stated is None:
            continue  # A per-store sheet labels its total row with the store name.
        if not close_enough(data_sum, stated):
            raise Rejected(
                [
                    f"SALES_EXPORT_INCONSISTENT: {path.name}[{name}] rows sum to {data_sum} "
                    f"but the export states {stated}; the export is truncated or edited"
                ]
            )
        window = declared_window(rows)
        window_start, cutoff = window if window else (None, None)
        return data_sum, f"{name}:{header_name}", window_start, cutoff

    if seen_column:
        raise Rejected(
            [f"{path.name}: no sheet states a total row, so the export cannot be checked"]
        )
    raise Rejected(
        [f"{path.name}: no gross-total column; expected one of {list(D05_TOTAL_HEADERS)}"]
    )


def build_verification(
    revenue: list[dict[str, Any]], stream: str, external: float, source_id: str
) -> dict[str, Any]:
    """Compare one revenue stream against an independent external total."""
    matches = [r for r in revenue if r["breakdown"] == "STREAM" and r["stream"] == stream]
    if not matches:
        available = sorted(
            r["stream"] for r in revenue if r["breakdown"] == "STREAM"
        )
        raise Rejected([f"no stream named {stream!r}; the workbook has {available}"])
    authoritative = matches[0]["amount_incl_vat"]
    difference = authoritative - external
    return {
        "stream": stream,
        "authoritative_amount": authoritative,
        "external_source_id": source_id,
        "external_amount": external,
        "difference": difference,
        "status": "MATCHED" if close_enough(difference, 0.0) else "DIVERGED",
    }


def close_enough(left: float, right: float) -> bool:
    """Compare two money figures, tolerating float noise but nothing larger."""
    return abs(left - right) < 0.005


def validate(
    revenue: list[dict[str, Any]],
    expenses: list[dict[str, Any]],
    stated_profit: float | None,
) -> list[str]:
    """Return the identities that failed. An empty list means the period is acceptable."""
    failures: list[str] = []

    def revenue_total(breakdown: str, scope: str | None = None) -> float:
        return sum(
            r["amount_incl_vat"]
            for r in revenue
            if r["breakdown"] == breakdown and r["scope"] == scope
        )

    total = revenue_total("TOTAL")
    for name, rule in (
        ("REVENUE_CATEGORY_SUM", revenue_total("CATEGORY")),
        ("REVENUE_STREAM_SUM", revenue_total("STREAM")),
    ):
        if not close_enough(rule, total):
            failures.append(f"{name}: {rule} != TOTAL {total}")

    weeks = sum(r["amount_incl_vat"] for r in revenue if r["breakdown"] == "WEEK")
    if not close_enough(weeks, total):
        failures.append(f"REVENUE_WEEK_SUM: {weeks} != TOTAL {total}")

    for record in revenue:
        if record["amount_ex_vat"] is None or record["output_vat"] is None:
            continue
        if record["amount_ex_vat"] < 0 or record["output_vat"] < 0:
            failures.append(f"REVENUE_NEGATIVE: {record['breakdown']} {record['scope']}")
        parts = record["amount_ex_vat"] + record["output_vat"]
        if not close_enough(parts, record["amount_incl_vat"]):
            failures.append(
                f"REVENUE_VAT_SUM: {record['breakdown']} {record['category'] or ''} "
                f"{parts} != {record['amount_incl_vat']}"
            )

    scopes = {r["scope"] for r in expenses}
    for scope in scopes:
        rows = [r for r in expenses if r["scope"] == scope]
        expense_total = sum(r["amount"] for r in rows if r["breakdown"] == "TOTAL")
        groups = [r for r in rows if r["breakdown"] == "GROUP"]
        group_sum = sum(r["amount"] for r in groups)
        if not close_enough(group_sum, expense_total):
            failures.append(
                f"EXPENSE_GROUP_SUM[{scope or 'period'}]: {group_sum} != TOTAL {expense_total}"
            )
        for group in groups:
            lines = [
                r
                for r in rows
                if r["breakdown"] == "LINE" and r["group_code"] == group["group_code"]
            ]
            if not lines:
                continue  # A group may be stated as a single figure with no lines.
            line_sum = sum(r["amount"] for r in lines)
            if not close_enough(line_sum, group["amount"]):
                failures.append(
                    f"EXPENSE_LINE_SUM[{scope or 'period'}/{group['group_code']}]: "
                    f"{line_sum} != GROUP {group['amount']}"
                )

    if stated_profit is not None:
        period_expense = sum(
            r["amount"] for r in expenses if r["breakdown"] == "TOTAL" and r["scope"] is None
        )
        derived = total - period_expense
        if not close_enough(derived, stated_profit):
            failures.append(f"DERIVED_PROFIT: {derived} != stated {stated_profit}")

    return failures


def week_labels(rows: list[list[Any]], revenue_row: int) -> list[str]:
    header = rows[revenue_row - 1] if revenue_row else []
    labels = []
    for index, column in enumerate(WEEK_COLS, start=1):
        value = header[column] if column < len(header) else None
        labels.append(value.strip() if isinstance(value, str) else f"Week {index}")
    return labels


def parse(path: Path, args: argparse.Namespace) -> dict[str, Any]:
    rows = read_pl(path)
    revenue_row, expense_row, profit_row = section_bounds(rows)
    weeks = week_labels(rows, revenue_row)

    revenue = build_revenue(rows, revenue_row, expense_row, weeks)
    expenses = build_expenses(rows, expense_row, profit_row, weeks)
    stated_profit = amount(rows[profit_row], MONTH_COL)

    failures = validate(revenue, expenses, stated_profit)
    if failures:
        raise Rejected(failures)

    # The workbook states no cutoff of its own: it aggregates streams whose cutoffs
    # differ, so a null here records that the cutoff is unknown for this document.
    authoritative_id = "SRC-1"
    sources = [
        {
            "source_id": authoritative_id,
            "role": AUTHORITATIVE,
            "label": args.source_label or path.name,
            "sha256": sha256_file(path),
            "stream": None,
            "business_day_cutoff": None,
            "window_start": None,
        }
    ]

    verifications = []
    if args.verify_sales:
        external, located, window_start, cutoff = read_sales_total(
            args.verify_sales, args.verify_sheet
        )
        if window_start is not None and window_start != args.period_start:
            raise Rejected(
                [
                    f"SOURCE_WINDOW_MISMATCH: {args.verify_sales.name} declares a window "
                    f"starting {window_start}, but the period starts {args.period_start}"
                ]
            )
        verification_id = "SRC-2"
        sources.append(
            {
                "source_id": verification_id,
                "role": VERIFICATION,
                "label": f"{args.verify_sales.name}[{located}]",
                "sha256": sha256_file(args.verify_sales),
                "stream": args.verify_stream,
                "business_day_cutoff": cutoff,
                "window_start": window_start,
            }
        )
        verifications.append(
            build_verification(revenue, args.verify_stream, external, verification_id)
        )

    for record in revenue:
        record["source_id"] = authoritative_id
    for record in expenses:
        record["source_id"] = authoritative_id

    return {
        "client_id": args.client_id,
        "project_id": args.project_id,
        "period_start": args.period_start,
        "period_end": args.period_end,
        "currency": args.currency,
        "sources": sources,
        "revenue_records": revenue,
        "expense_records": expenses,
        "verifications": verifications,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--client-id", required=True)
    parser.add_argument("--project-id", required=True)
    parser.add_argument("--period-start", required=True, help="YYYY-MM-DD business date")
    parser.add_argument("--period-end", required=True, help="YYYY-MM-DD business date")
    parser.add_argument("--currency", required=True)
    parser.add_argument("--source-label", default=None)
    parser.add_argument(
        "--verify-sales",
        type=Path,
        help="a POS sales export to verify one revenue stream against",
    )
    parser.add_argument(
        "--verify-stream",
        default="2F",
        help="the stream the sales export covers (default: %(default)s)",
    )
    parser.add_argument(
        "--verify-sheet",
        default=None,
        help="the sheet of the sales export to read; by default the first usable one",
    )
    parser.add_argument("--out", type=Path, help="write the payload here instead of stdout")
    args = parser.parse_args(argv)

    for name in ("period_start", "period_end"):
        try:
            dt.date.fromisoformat(getattr(args, name))
        except ValueError:
            print(f"rejected: --{name.replace('_', '-')} is not an ISO date", file=sys.stderr)
            return 2

    try:
        payload = parse(args.workbook, args)
    except Rejected as rejection:
        print("rejected: the workbook failed validation, nothing was written", file=sys.stderr)
        for failure in rejection.failures:
            print(f"  {failure}", file=sys.stderr)
        return 1

    for verification in payload["verifications"]:
        if verification["status"] == "DIVERGED":
            print(
                f"warning: stream {verification['stream']} differs from "
                f"{verification['external_source_id']} by {verification['difference']:,.0f}; "
                "recorded as DIVERGED for review, the period is still accepted",
                file=sys.stderr,
            )

    text = json.dumps(payload, indent=2, ensure_ascii=False)
    if args.out:
        args.out.write_text(text + "\n", encoding="utf-8")
        statuses = ", ".join(
            f"{v['stream']} {v['status']}" for v in payload["verifications"]
        )
        print(
            f"accepted: {len(payload['revenue_records'])} revenue and "
            f"{len(payload['expense_records'])} expense records -> {args.out}"
            + (f" [{statuses}]" if statuses else "")
        )
    else:
        print(text)
    return 0


if __name__ == "__main__":
    sys.exit(main())
