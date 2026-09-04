#!/usr/bin/env python3
"""Dump the field structure of a FABi report export (.xlsx / .csv).

Structure only: sheet names, dimensions, the detected header row, each column's
inferred type, and a small sample. Amounts are shown so types can be judged;
this prints to the terminal and writes nothing.
"""
import csv
import signal
import sys
from pathlib import Path


def infer(values):
    seen = set()
    for v in values:
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        seen.add(type(v).__name__)
    return "/".join(sorted(seen)) or "empty"


def header_row(rows, limit=25):
    """First row whose non-empty cells are mostly text -> likely the header."""
    best, best_score = 0, -1
    for i, row in enumerate(rows[:limit]):
        cells = [c for c in row if c is not None and str(c).strip()]
        if len(cells) < 2:
            continue
        score = sum(isinstance(c, str) for c in cells) * len(cells)
        if score > best_score:
            best, best_score = i, score
    return best


def dump(name, rows):
    print(f"\n=== {name} ===")
    print(f"rows={len(rows)} cols={max((len(r) for r in rows), default=0)}")
    if not rows:
        return
    h = header_row(rows)
    print(f"header row index: {h}")
    for i, row in enumerate(rows[: h + 1]):
        cells = [str(c) for c in row if c is not None and str(c).strip()]
        print(f"  preamble[{i}]: {' | '.join(cells)[:200]}")
    headers = [str(c).strip() if c is not None else "" for c in rows[h]]
    body = rows[h + 1 :]
    print(f"\n{'#':>3}  {'header':<38}  {'type':<18}  sample")
    for c, name_ in enumerate(headers):
        col = [r[c] if c < len(r) else None for r in body]
        sample = next(
            (repr(v)[:40] for v in col if v is not None and str(v).strip()), ""
        )
        print(f"{c:>3}  {name_[:38]:<38}  {infer(col[:400]):<18}  {sample}")
    print(f"\ndata rows: {len(body)}")


def main(path):
    p = Path(path)
    if p.suffix.lower() in {".xlsx", ".xlsm"}:
        import openpyxl

        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        print(f"file: {p.name}  sheets: {wb.sheetnames}")
        for ws in wb.worksheets:
            dump(ws.title, [list(r) for r in ws.iter_rows(values_only=True)])
    else:
        raw = p.read_bytes()
        for enc in ("utf-8-sig", "utf-8", "cp1258", "latin-1"):
            try:
                text = raw.decode(enc)
                print(f"file: {p.name}  encoding: {enc}")
                break
            except UnicodeDecodeError:
                continue
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
        print(f"delimiter: {dialect.delimiter!r}")
        dump(p.name, list(csv.reader(text.splitlines(), dialect)))


if __name__ == "__main__":
    # Stay quiet when the output is truncated by a pager such as `head`.
    signal.signal(signal.SIGPIPE, signal.SIG_DFL)
    if len(sys.argv) < 2:
        sys.exit("usage: inspect_report.py <file.xlsx|file.csv> ...")
    for a in sys.argv[1:]:
        main(a)
